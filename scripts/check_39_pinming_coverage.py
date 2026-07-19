# -*- coding: utf-8 -*-
"""39品名すべてにKSデータが分類されているか."""
from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd

SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
HIER = Path(
    r"E:\KSS-KSファイル集計\output\reports\KS_指定12項目抽出_分類階層_20260718_153724.xlsx"
)
EXCL = Path(r"E:\sales-dashboard-2\data\ks-caiwu-excess-exclusions.json")
SALES = Path(r"E:\sales-dashboard-2")
BATCH = SALES / "scripts" / "classify-pinming-batch.ts"
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
SKIP = re.compile(r"小\s*计|合\s*计|品\s*名|出口产品")
PAINT = re.compile(r"来料喷涂|喷涂加工|運費|运费")


def sakin_names() -> list[str]:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    names: set[str] = set()
    for sh in MONTHS:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        for r in range(4, (ws.max_row or 0) + 1):
            a = ws.cell(r, 1).value
            d = ws.cell(r, 4).value
            if not a or SKIP.search(str(a)):
                continue
            if isinstance(d, (int, float)) and float(d) != 0:
                names.add(str(a).strip())
    wb.close()
    return sorted(names)


def classify_ks() -> pd.DataFrame:
    k = pd.read_excel(HIER, sheet_name="明细")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["month"] = k["日期"].dt.strftime("%Y-%m")
    k = k[k["month"].isin(MONTHS)].copy()
    k = k[~k["description"].astype(str).str.contains(PAINT, na=False)]
    if EXCL.exists():
        data = json.loads(EXCL.read_text(encoding="utf-8"))
        k = k[~k.index.isin(set(data.get("excluded_row_indices", [])))]
        keys = {
            (
                str(l.get("ks_doc_no", "")).upper(),
                str(l.get("delivery_list_no", "")).upper(),
                str(l.get("description", ""))[:80],
                round(float(l.get("amount_hkd", 0)), 2),
            )
            for l in data.get("lines", [])
        }
        k = k[
            ~k.apply(
                lambda r: (
                    str(r.get("單號", "")).upper(),
                    str(r.get("出貨清單號", "")).upper(),
                    str(r.get("description", ""))[:80],
                    round(float(r.get("Amount(HKD)", 0)), 2),
                )
                in keys,
                axis=1,
            )
        ]
    req = []
    for i, r in k.iterrows():
        req.append(
            {
                "id": str(i),
                "product_name": None if pd.isna(r.get("品名")) else r.get("品名"),
                "description": None if pd.isna(r.get("description")) else r.get("description"),
                "mid_category": None if pd.isna(r.get("材料中分类")) else r.get("材料中分类"),
                "product_code": None if pd.isna(r.get("产品代码")) else r.get("产品代码"),
            }
        )
    tmp = Path(r"E:\factory_monitoring_system\logs\_tmp_39check")
    tmp.mkdir(exist_ok=True)
    inp, outp = tmp / "in.json", tmp / "out.json"
    inp.write_text(json.dumps(req, ensure_ascii=False), encoding="utf-8")
    exe = "npx.cmd" if os.name == "nt" else "npx"
    proc = subprocess.run(
        [exe, "--yes", "tsx", str(BATCH), str(inp), str(outp)],
        cwd=str(SALES),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        shell=os.name == "nt",
    )
    if proc.returncode != 0:
        raise SystemExit(proc.stderr or proc.stdout)
    mp = {x["id"]: x["pinming"] for x in json.loads(outp.read_text(encoding="utf-8"))}
    k = k.copy()
    k["pinming"] = [mp[str(i)] for i in k.index]
    return k


def main() -> None:
    names39 = sakin_names()
    k = classify_ks()
    ks_by = k.groupby("pinming")["Amount(HKD)"].sum().sort_values(ascending=False)

    print("=== 質問: 39種類ですべてのデータを分けたか ===")
    print(f"佐近Excel 1-6月 品名数: {len(names39)}")
    print(f"KS明細行数: {len(k)}")
    print(f"KS分類先ユニーク数: {k['pinming'].nunique()}")
    print()

    # 39 names with KS amount
    has_ks = []
    no_ks = []
    for n in names39:
        amt = float(ks_by.get(n, 0))
        if amt > 0:
            has_ks.append((n, amt))
        else:
            no_ks.append(n)

    print(f"39品名のうち KS側に1円以上入っている: {len(has_ks)}")
    print(f"39品名のうち KS側が 0（未分類/別名統合）: {len(no_ks)}")
    print()
    print("--- KS側 0 の品名（39のうち）---")
    for n in no_ks:
        print(f"  {n}")

    print()
    print("--- KS分類先（39以外・未分類含む）---")
    outside = [(p, a) for p, a in ks_by.items() if p not in names39 and a > 0]
    for p, a in sorted(outside, key=lambda x: -x[1]):
        print(f"  {p}: {a:,.0f}")

    unclassified = float(k[k["pinming"] == "未分類"]["Amount(HKD)"].sum())
    total = float(k["Amount(HKD)"].sum())
    in39 = float(k[k["pinming"].isin(names39)]["Amount(HKD)"].sum())
    print()
    print(f"KS総額: {total:,.0f}")
    print(f"39品名に直接入った額: {in39:,.0f} ({in39/total*100:.1f}%)")
    print(f"39以外+未分類: {total-in39:,.0f}")
    print(f"未分類のみ: {unclassified:,.0f}")


if __name__ == "__main__":
    main()
