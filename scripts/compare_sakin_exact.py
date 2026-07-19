# -*- coding: utf-8 -*-
"""佐近 Excel 品名×月 vs 現在 KS 分類 — 完全一致か."""
from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd

SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
HIER = Path(
    r"E:\KSS-KSファイル集計\output\reports\KS_指定12項目抽出_分類階層_20260718_153724.xlsx"
)
EXCL = Path(r"E:\sales-dashboard-2\data\ks-caiwu-excess-exclusions.json")
SALES = Path(r"E:\sales-dashboard-2")
BATCH = SALES / "scripts" / "classify-pinming-batch.ts"
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
SKIP = re.compile(r"小\s*计|合\s*计|品\s*名|出口产品")
PAINT = re.compile(r"来料喷涂|喷涂加工|運費|运费")


def sakin_matrix() -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    out: dict[str, dict[str, float]] = {}
    for sh in wb.sheetnames:
        if sh not in MONTHS:
            continue
        ws = wb[sh]
        for r in range(4, (ws.max_row or 0) + 1):
            a = ws.cell(r, 1).value
            d = ws.cell(r, 4).value
            if not a or SKIP.search(str(a)):
                continue
            name = str(a).strip()
            if isinstance(d, (int, float)):
                out.setdefault(name, {})
                out[name][sh] = out[name].get(sh, 0) + float(d)
    wb.close()
    return out


def classify_ks() -> pd.DataFrame:
    k = pd.read_excel(HIER, sheet_name="明细")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["month"] = k["日期"].dt.strftime("%Y-%m")
    k = k[k["month"].isin(MONTHS)].copy()
    k = k[~k["description"].astype(str).str.contains(PAINT, na=False)]
    if EXCL.exists():
        data = json.loads(EXCL.read_text(encoding="utf-8"))
        k = k[~k.index.isin(set(data.get("excluded_row_indices", [])))]
        keys = {
            (
                str(l.get("ks_doc_no", "")).upper(),
                str(l.get("delivery_list_no", "")).upper(),
                str(l.get("description", ""))[:80],
                round(float(l.get("amount_hkd", 0)), 2),
            )
            for l in data.get("lines", [])
        }
        k = k[
            ~k.apply(
                lambda r: (
                    str(r.get("單號", "")).upper(),
                    str(r.get("出貨清單號", "")).upper(),
                    str(r.get("description", ""))[:80],
                    round(float(r.get("Amount(HKD)", 0)), 2),
                )
                in keys,
                axis=1,
            )
        ]
    req = []
    for i, r in k.iterrows():
        req.append(
            {
                "id": str(i),
                "product_name": None if pd.isna(r.get("品名")) else r.get("品名"),
                "description": None if pd.isna(r.get("description")) else r.get("description"),
                "mid_category": None if pd.isna(r.get("材料中分类")) else r.get("材料中分类"),
                "product_code": None if pd.isna(r.get("产品代码")) else r.get("产品代码"),
            }
        )
    tmp = Path(r"E:\factory_monitoring_system\logs\_tmp_exact")
    tmp.mkdir(exist_ok=True)
    inp, outp = tmp / "in.json", tmp / "out.json"
    inp.write_text(json.dumps(req, ensure_ascii=False), encoding="utf-8")
    exe = "npx.cmd" if os.name == "nt" else "npx"
    proc = subprocess.run(
        [exe, "--yes", "tsx", str(BATCH), str(inp), str(outp)],
        cwd=str(SALES),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        shell=os.name == "nt",
    )
    if proc.returncode != 0:
        raise SystemExit(proc.stderr or proc.stdout)
    mp = {x["id"]: x["pinming"] for x in json.loads(outp.read_text(encoding="utf-8"))}
    k = k.copy()
    k["pinming"] = [mp[str(i)] for i in k.index]
    return k


def main() -> None:
    sakin = sakin_matrix()
    k = classify_ks()
    ks: dict[str, dict[str, float]] = {}
    for _, r in k.iterrows():
        ks.setdefault(r["pinming"], {})
        ks[r["pinming"]][r["month"]] = ks[r["pinming"]].get(r["month"], 0) + float(
            r["Amount(HKD)"]
        )

    all_names = set(sakin) | set(ks)
    mismatch_cells = 0
    total_cells = 0
    exact_cells = 0
    rows: list[tuple] = []

    for name in sorted(all_names):
        for m in MONTHS:
            sk = sakin.get(name, {}).get(m, 0)
            kk = ks.get(name, {}).get(m, 0)
            total_cells += 1
            if abs(sk - kk) < 1:
                exact_cells += 1
            else:
                mismatch_cells += 1
                if sk or kk:
                    rows.append((name, m, sk, kk, kk - sk))

    print("=== 佐近 Excel vs 現在 KS 分類 ===")
    print(f"品名×月セル数: {total_cells}")
    print(f"完全一致(±1HKD): {exact_cells}")
    print(f"不一致: {mismatch_cells}")
    print(f"一致率: {exact_cells / total_cells * 100:.1f}%")
    print()
    print("--- 不一致セル (金額あり) 上位15 ---")
    rows.sort(key=lambda x: abs(x[4]), reverse=True)
    for r in rows[:15]:
        print(f"  {r[0]} | {r[1]} | 佐近 {r[2]:,.0f} | KS {r[3]:,.0f} | 差 {r[4]:+,.0f}")

    print()
    print("--- 月次合計 ---")
    for m in MONTHS:
        sk = sum(v.get(m, 0) for v in sakin.values())
        kk = float(k[k["month"] == m]["Amount(HKD)"].sum())
        print(f"  {m} 佐近 {sk:,.0f}  KS {kk:,.0f}  差 {kk - sk:+,.0f}")

    # names only in one side
    only_sakin = sorted(set(sakin) - set(ks))
    only_ks = sorted(set(ks) - set(sakin))
    print()
    print(f"佐近にのみ存在(金額>0): {len(only_sakin)}")
    for n in only_sakin[:10]:
        tot = sum(sakin[n].values())
        if tot:
            print(f"  {n}: {tot:,.0f}")
    print(f"KS分類にのみ存在(金額>0): {len(only_ks)}")
    for n in only_ks[:10]:
        tot = sum(ks[n].values())
        if tot:
            print(f"  {n}: {tot:,.0f}")


if __name__ == "__main__":
    main()
