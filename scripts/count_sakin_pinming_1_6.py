# -*- coding: utf-8 -*-
"""佐近 Excel 1-6月 品名（分類）の総数と月別出現."""
from __future__ import annotations

import re
from collections import Counter
from pathlib import Path

import openpyxl

SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
SKIP = re.compile(r"小\s*计|合\s*计|品\s*名|出口产品")


def main() -> None:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    by_name: dict[str, dict[str, float]] = {}

    for sh in MONTHS:
        if sh not in wb.sheetnames:
            print("missing sheet", sh)
            continue
        ws = wb[sh]
        for r in range(4, (ws.max_row or 0) + 1):
            a = ws.cell(r, 1).value
            d = ws.cell(r, 4).value
            if not a or SKIP.search(str(a)):
                continue
            name = str(a).strip()
            amt = float(d) if isinstance(d, (int, float)) else 0.0
            rec = by_name.setdefault(name, {m: 0.0 for m in MONTHS})
            rec[sh] = rec.get(sh, 0) + amt

    wb.close()

    active = {n: v for n, v in by_name.items() if sum(v.values()) > 0}

    print("=== 佐近 Excel 1-6月 品名（分類）===")
    print(f"全ユニーク品名数（1-6月で金額>0が1回以上）: {len(active)}")
    print()

    month_counts: Counter[int] = Counter()
    for v in active.values():
        month_counts[sum(1 for m in MONTHS if v.get(m, 0) > 0)] += 1

    print("--- 金額がある月数 ---")
    for c in sorted(month_counts):
        print(f"  {c}ヶ月: {month_counts[c]}品名")

    print()
    print("--- 全品名一覧 ---")
    for i, n in enumerate(sorted(active.keys()), 1):
        vals = [active[n].get(m, 0) for m in MONTHS]
        months_present = [m.split("-")[1] + "月" for m, v in zip(MONTHS, vals) if v > 0]
        total = sum(vals)
        present = " ".join(months_present) if months_present else "なし"
        print(f"{i:2}. {n} | 計 {total:,.0f} | {present}")

    all6 = [n for n, v in active.items() if all(v.get(m, 0) > 0 for m in MONTHS)]
    print()
    print(f"--- 1-6月 毎月すべて金額あり: {len(all6)}品名 ---")
    for n in sorted(all6):
        print(f"  {n}")

    partial = [n for n, v in active.items() if 0 < sum(1 for m in MONTHS if v.get(m, 0) > 0) < 6]
    print()
    print(f"--- 一部の月だけ金額あり: {len(partial)}品名 ---")


if __name__ == "__main__":
    main()
