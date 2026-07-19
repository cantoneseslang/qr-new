# -*- coding: utf-8 -*-
"""
CAIWU ↔ KS 明細級 左右対比 Excel。

左 CAIWU: 客户名称, 项目名称, 订货单号, sales, 产品名称, 产品尺寸,
         数量, 单位, 每件, 每件净重量, 总净重量, 单价, 金额, 出货清单号
右 KS:    客户名称, 单号, 日期, 本司订单编号, 项目名称, 销售人员,
         出货清单号, Description, Qty, UnitPrice, Amount

突合: 出貨清單號優先（スペース分割・R接尾辞吸収）。同一出荷の全明細を左右に並べる。

Usage:
  py -3 scripts/export_caiwu_ks_line_detail.py
  py -3 scripts/export_caiwu_ks_line_detail.py --from-month 2026-01 --to-month 2026-05
"""
from __future__ import annotations

import argparse
import json
import os
import re
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

REP = Path(r"E:\KSS-KSファイル集計\output\reports")
OUT_DIR = Path(r"E:\factory_monitoring_system\logs")
SALES_DASHBOARD = Path(r"E:\sales-dashboard-2")
PINMING_BATCH = SALES_DASHBOARD / "scripts" / "classify-pinming-batch.ts"


def classify_pinming_batch(
    rows: list[dict],
    tmp_dir: Path,
) -> dict[str, str]:
    """Call dashboard classifyCaiwuProductName via tsx. rows need id + name fields."""
    if not rows:
        return {}
    tmp_dir.mkdir(parents=True, exist_ok=True)
    in_path = tmp_dir / "_pinming_in.json"
    out_path = tmp_dir / "_pinming_out.json"
    # sanitize NaN/NA for JSON
    clean_rows = []
    for r in rows:
        clean = {}
        for k, v in r.items():
            if v is None or (isinstance(v, float) and pd.isna(v)):
                clean[k] = None
            elif isinstance(v, float) and (v != v):  # NaN
                clean[k] = None
            else:
                try:
                    if pd.isna(v):
                        clean[k] = None
                    else:
                        clean[k] = v
                except (TypeError, ValueError):
                    clean[k] = v
        clean_rows.append(clean)
    in_path.write_text(
        json.dumps(clean_rows, ensure_ascii=False, allow_nan=False),
        encoding="utf-8",
    )
    cmd = [
        "npx",
        "--yes",
        "tsx",
        str(PINMING_BATCH),
        str(in_path),
        str(out_path),
    ]
    # Windows npx.cmd
    exe = "npx.cmd" if os.name == "nt" else "npx"
    cmd[0] = exe
    proc = subprocess.run(
        cmd,
        cwd=str(SALES_DASHBOARD),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if proc.returncode != 0:
        raise SystemExit(
            f"pinming classify failed:\n{proc.stdout}\n{proc.stderr}"
        )
    data = json.loads(out_path.read_text(encoding="utf-8"))
    return {str(r["id"]): str(r["pinming"]) for r in data}


def find_latest_caiwu(rep: Path) -> Path:
    cands = sorted(
        [p for p in rep.glob("CAIWU_货值清单总表抽出_*.csv") if "差分更新" not in p.name],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not cands:
        raise SystemExit("CAIWU CSV not found")
    return cands[0]


def find_latest_ks(rep: Path) -> Path:
    skip = ("分類分析", "分類階層", "大分類と詳細", "大中小分類", "产品代码")
    cands = sorted(
        [
            p
            for p in rep.glob("KS_指定12項目抽出_*.csv")
            if all(s not in p.name for s in skip)
        ],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not cands:
        raise SystemExit("KS CSV not found")
    return cands[0]


def norm_key(v: object) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ""
    return str(v).strip().upper().replace(" ", "").replace("　", "")


def split_ship_keys(v: object) -> list[str]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return []
    s = str(v).strip()
    if not s:
        return []
    parts = [norm_key(p) for p in re.split(r"[;；,/|\s]+", s)]
    return [p for p in parts if p]


def ship_match_keys(v: str) -> list[str]:
    base = norm_key(v)
    if not base:
        return []
    keys = [base]
    stripped = re.sub(r"R\d+$", "", base)
    if stripped and stripped != base:
        keys.append(stripped)
    return keys


def canonical_ship(ship: str) -> str:
    keys = ship_match_keys(ship)
    if not keys:
        return ""
    return min(keys, key=len)


def uniq_join(vals) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for v in vals:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if not s or s.lower() == "nan":
            continue
        if s not in seen:
            seen.add(s)
            out.append(s)
    return " / ".join(out)


def main() -> None:
    parser = argparse.ArgumentParser(description="CAIWU↔KS line-level side-by-side")
    parser.add_argument("--caiwu", type=Path, default=None)
    parser.add_argument("--ks", type=Path, default=None)
    parser.add_argument("--from-month", default="2026-01")
    parser.add_argument("--to-month", default="2026-05")
    parser.add_argument("--out-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    caiwu_path = args.caiwu or find_latest_caiwu(REP)
    ks_path = args.ks or find_latest_ks(REP)
    print(f"caiwu: {caiwu_path}")
    print(f"ks: {ks_path}")

    c = pd.read_csv(caiwu_path, encoding="utf-8-sig")
    k = pd.read_csv(ks_path, encoding="utf-8-sig")

    for col in ["amount_hkd", "qty", "unit_price", "pack_count", "unit_net_weight", "total_net_weight"]:
        if col in c.columns:
            c[col] = pd.to_numeric(c[col], errors="coerce")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["Qty"] = pd.to_numeric(k["Qty"], errors="coerce")
    k["(HKD)unit price"] = pd.to_numeric(k["(HKD)unit price"], errors="coerce")
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["sale_month"] = k["日期"].dt.strftime("%Y-%m")
    c["sale_month"] = c["sale_month"].astype(str)
    c["shipment_date"] = pd.to_datetime(c["shipment_date"], errors="coerce")

    c = c[
        (c["sale_month"] >= args.from_month) & (c["sale_month"] <= args.to_month)
    ].copy()
    k = k[
        (k["sale_month"] >= args.from_month) & (k["sale_month"] <= args.to_month)
    ].copy()

    c = c.reset_index(drop=True)
    c["_cid"] = range(len(c))
    k = k.reset_index(drop=True)
    k["_kid"] = range(len(k))

    # 佐近品名（dashboard classifyCaiwuProductName と同じ）
    print("classifying pinming (CAIWU+KS)...", flush=True)
    pinming_req = []
    for _, r in c.iterrows():
        pinming_req.append(
            {
                "id": f"C{int(r['_cid'])}",
                "product_name": r.get("product_name"),
                "product_size": r.get("product_size"),
            }
        )
    for _, r in k.iterrows():
        pinming_req.append(
            {
                "id": f"K{int(r['_kid'])}",
                "product_name": r.get("品名"),
                "product_size": None,
                "description": r.get("description"),
                "mid_category": r.get("材料中分类"),
                "product_code": r.get("产品代码"),
            }
        )
    pinming_map = classify_pinming_batch(pinming_req, args.out_dir / "_tmp_pinming")
    c["品名"] = c["_cid"].map(lambda i: pinming_map.get(f"C{int(i)}", "未分類"))
    k["品名"] = k["_kid"].map(lambda i: pinming_map.get(f"K{int(i)}", "未分類"))
    print(
        f"  CAIWU pinming sample: {c['品名'].value_counts().head(5).to_dict()}",
        flush=True,
    )
    print(
        f"  KS pinming sample: {k['品名'].value_counts().head(5).to_dict()}",
        flush=True,
    )

    # Exact ship token only（R接尾辞は別キー。左右の出货清单号を同一行で一致させる）
    c_exact: dict[str, list[int]] = {}
    for _, r in c.iterrows():
        ships = split_ship_keys(r.get("delivery_list_no"))
        if not ships:
            continue
        for ship in ships:
            c_exact.setdefault(ship, []).append(int(r["_cid"]))

    k_exact: dict[str, list[int]] = {}
    for _, r in k.iterrows():
        ships = split_ship_keys(r.get("出貨清單號"))
        if not ships:
            continue
        for ship in ships:
            k_exact.setdefault(ship, []).append(int(r["_kid"]))

    used_c: set[int] = set()
    used_k: set[int] = set()
    groups: list[dict] = []

    def take_cids(ship: str) -> list[int]:
        return [i for i in c_exact.get(ship, []) if i not in used_c]

    def take_kids(ship: str) -> list[int]:
        # 同一KS行が複数出荷トークンを持つ場合、先に使われたら再利用しない
        return [i for i in k_exact.get(ship, []) if i not in used_k]

    # 1) exact 出货清单号
    for ship in sorted(set(c_exact) | set(k_exact)):
        if not ship:
            continue
        cids = take_cids(ship)
        kids = take_kids(ship)
        if not cids and not kids:
            continue
        if cids and kids:
            c_orders = {
                norm_key(x)
                for x in c.loc[c["_cid"].isin(cids), "order_no"].tolist()
                if norm_key(x)
            }
            k_orders = {
                norm_key(x)
                for x in k.loc[k["_kid"].isin(kids), "本司訂單編號"].tolist()
                if norm_key(x)
            }
            how = "exact" if c_orders and k_orders and c_orders == k_orders else "ship"
            status = None
        elif cids:
            how, status = "caiwu_only", "CAIWUのみ（KS欠）"
        else:
            how, status = "ks_only", "KSのみ（CAIWU欠）"
        used_c.update(cids)
        used_k.update(kids)
        groups.append(
            {
                "how": how,
                "status": status,
                "cids": sorted(set(cids)),
                "kids": sorted(set(kids)),
                "ship": ship,
            }
        )

    # 2) R接尾辞のゆるい突合（残ったもの同士）
    rem_c_ships: dict[str, list[int]] = {}
    for ship, ids in c_exact.items():
        left = [i for i in ids if i not in used_c]
        if left:
            rem_c_ships[ship] = left
    rem_k_ships: dict[str, list[int]] = {}
    for ship, ids in k_exact.items():
        left = [i for i in ids if i not in used_k]
        if left:
            rem_k_ships[ship] = left

    used_soft_c: set[str] = set()
    used_soft_k: set[str] = set()
    for c_ship, cids0 in sorted(rem_c_ships.items()):
        if c_ship in used_soft_c:
            continue
        cids = [i for i in cids0 if i not in used_c]
        if not cids:
            continue
        canon = canonical_ship(c_ship)
        # find KS ships with same canonical
        k_ships = [
            ks
            for ks in rem_k_ships
            if ks not in used_soft_k and canonical_ship(ks) == canon and ks != c_ship
        ]
        if not k_ships:
            continue
        kids: list[int] = []
        for ks in k_ships:
            for kid in rem_k_ships[ks]:
                if kid not in used_k and kid not in kids:
                    kids.append(kid)
            used_soft_k.add(ks)
        if not kids:
            continue
        used_soft_c.add(c_ship)
        used_c.update(cids)
        used_k.update(kids)
        groups.append(
            {
                "how": "ship_r",
                "status": None,
                "cids": sorted(set(cids)),
                "kids": sorted(set(kids)),
                "ship": f"{c_ship}≈{'/'.join(k_ships)}",
                "c_ship": c_ship,
                "k_ships": k_ships,
            }
        )

    # 3) remaining CAIWU (empty ship → order; else KS欠)
    for _, r in c.iterrows():
        cid = int(r["_cid"])
        if cid in used_c:
            continue
        ships = split_ship_keys(r.get("delivery_list_no"))
        if ships:
            used_c.add(cid)
            groups.append(
                {
                    "how": "caiwu_only",
                    "status": "CAIWUのみ（KS欠）",
                    "cids": [cid],
                    "kids": [],
                    "ship": ships[0],
                }
            )
            continue
        ok = norm_key(r.get("order_no"))
        kids = []
        if ok:
            for _, kr in k.iterrows():
                kid = int(kr["_kid"])
                if kid in used_k:
                    continue
                if norm_key(kr.get("本司訂單編號")) == ok:
                    kids.append(kid)
        if kids:
            used_c.add(cid)
            used_k.update(kids)
            groups.append(
                {
                    "how": "order_only",
                    "status": None,
                    "cids": [cid],
                    "kids": sorted(set(kids)),
                    "ship": ok,
                }
            )
        else:
            used_c.add(cid)
            groups.append(
                {
                    "how": "caiwu_only",
                    "status": "CAIWUのみ（KS欠）",
                    "cids": [cid],
                    "kids": [],
                    "ship": ok or f"cid{cid}",
                }
            )

    # 4) remaining KS
    for _, kr in k.iterrows():
        kid = int(kr["_kid"])
        if kid in used_k:
            continue
        used_k.add(kid)
        ships = split_ship_keys(kr.get("出貨清單號"))
        groups.append(
            {
                "how": "ks_only",
                "status": "KSのみ（CAIWU欠）",
                "cids": [],
                "kids": [kid],
                "ship": ships[0] if ships else norm_key(kr.get("本司訂單編號")),
            }
        )

    def blank_c() -> dict:
        return {f"C_{x}": "" for x in [
            "品名", "客户名称", "项目名称", "订货单号", "sales", "产品名称", "产品尺寸",
            "数量", "单位", "每件", "每件净重量", "总净重量", "单价", "金额", "出货清单号",
        ]}

    def blank_k() -> dict:
        return {f"K_{x}": "" for x in [
            "品名", "客户名称", "单号", "日期", "本司订单编号", "项目名称", "销售人员",
            "出货清单号", "Description", "Qty", "UnitPrice", "Amount",
        ]}

    def row_c(r: pd.Series) -> dict:
        return {
            "C_品名": r.get("品名", "") or "",
            "C_客户名称": r.get("customer_name", "") or "",
            "C_项目名称": r.get("project", "") or "",
            "C_订货单号": r.get("order_no", "") or "",
            "C_sales": r.get("sales", "") or "",
            "C_产品名称": r.get("product_name", "") or "",
            "C_产品尺寸": r.get("product_size", "") or "",
            "C_数量": r.get("qty", ""),
            "C_单位": r.get("unit", "") or "",
            "C_每件": r.get("pack_count", ""),
            "C_每件净重量": r.get("unit_net_weight", ""),
            "C_总净重量": r.get("total_net_weight", ""),
            "C_单价": r.get("unit_price", ""),
            "C_金额": r.get("amount_hkd", ""),
            "C_出货清单号": r.get("delivery_list_no", "") or "",
        }

    def row_k(r: pd.Series) -> dict:
        dt = r.get("日期")
        if pd.notna(dt):
            try:
                dt_s = pd.Timestamp(dt).strftime("%Y-%m-%d")
            except Exception:
                dt_s = str(dt)
        else:
            dt_s = ""
        return {
            "K_品名": r.get("品名", "") or "",
            "K_客户名称": r.get("客戶名稱", "") or "",
            "K_单号": r.get("單號", "") or "",
            "K_日期": dt_s,
            "K_本司订单编号": r.get("本司訂單編號", "") or "",
            "K_项目名称": r.get("項目名稱", "") or "",
            "K_销售人员": r.get("销售人员", "") or "",
            "K_出货清单号": r.get("出貨清單號", "") or "",
            "K_Description": r.get("description", "") or "",
            "K_Qty": r.get("Qty", ""),
            "K_UnitPrice": r.get("(HKD)unit price", ""),
            "K_Amount": r.get("Amount(HKD)", ""),
        }

    out_rows: list[dict] = []
    for g in groups:
        cids = g["cids"]
        kids = g["kids"]
        c_sub = c[c["_cid"].isin(cids)].copy() if cids else c.iloc[0:0]
        k_sub = k[k["_kid"].isin(kids)].copy() if kids else k.iloc[0:0]
        if len(c_sub):
            c_sub = c_sub.sort_values(["shipment_date", "_cid"])
        if len(k_sub):
            k_sub = k_sub.sort_values(["日期", "_kid"])

        c_amt = float(pd.to_numeric(c_sub["amount_hkd"], errors="coerce").fillna(0).sum()) if len(c_sub) else 0.0
        k_amt = float(pd.to_numeric(k_sub["Amount(HKD)"], errors="coerce").fillna(0).sum()) if len(k_sub) else 0.0

        status = g.get("status")
        if status is None:
            if g["how"] == "caiwu_only":
                status = "CAIWUのみ（KS欠）"
            elif g["how"] == "ks_only":
                status = "KSのみ（CAIWU欠）"
            else:
                c_orders = {
                    norm_key(x) for x in c_sub["order_no"].tolist() if norm_key(x)
                }
                k_orders = {
                    norm_key(x) for x in k_sub["本司訂單編號"].tolist() if norm_key(x)
                }
                order_diff = c_orders != k_orders
                amt_ok = abs(c_amt - k_amt) <= 0.5
                if g["how"] == "ship_r":
                    status = "出荷番号R差（金額一致）" if amt_ok else "出荷番号R差（金額差）"
                elif order_diff and g["how"] in ("ship", "ship_only"):
                    status = "注文番号不一致（金額一致）" if amt_ok else "注文番号不一致（金額差）"
                elif amt_ok:
                    status = "一致"
                else:
                    status = "金額不一致"

        month = ""
        if len(c_sub):
            month = str(c_sub["sale_month"].iloc[0])
        elif len(k_sub):
            month = str(k_sub["sale_month"].iloc[0])

        # 左右の出货清单号を同じキーで揃える（exact）。R差は左右で実番号を出す。
        if g["how"] == "ship_r":
            c_ship_key = str(g.get("c_ship") or "")
            k_ship_key = "/".join(g.get("k_ships") or [])
        else:
            c_ship_key = str(g.get("ship") or "")
            k_ship_key = c_ship_key

        n = max(len(c_sub), len(k_sub), 1)
        c_recs = c_sub.to_dict("records") if len(c_sub) else []
        k_recs = k_sub.to_dict("records") if len(k_sub) else []

        for i in range(n):
            row: dict = {
                "状態": status if i == 0 else "",
                "月": month if i == 0 else "",
                "突合": g["how"] if i == 0 else "",
                "対比出货清单号": c_ship_key if g["how"] != "ship_r" else str(g.get("ship") or ""),
                "グループ金額差_KS-C": (k_amt - c_amt) if i == 0 else "",
                "グループC合計": c_amt if i == 0 else "",
                "グループK合計": k_amt if i == 0 else "",
            }
            if i < len(c_recs):
                crow = row_c(pd.Series(c_recs[i]))
                # exact突合時は左右とも対比キーで出货を揃える
                if g["how"] != "ship_r" and c_ship_key:
                    crow["C_出货清单号"] = c_ship_key
                row.update(crow)
            else:
                bc = blank_c()
                # 明細が無い側も出货キーだけ埋めて左右位置を揃える
                if c_ship_key and g["how"] not in ("ks_only",):
                    bc["C_出货清单号"] = c_ship_key
                elif g["how"] == "ks_only" and k_ship_key:
                    bc["C_出货清单号"] = k_ship_key
                row.update(bc)
            if i < len(k_recs):
                krow = row_k(pd.Series(k_recs[i]))
                if g["how"] == "ship_r":
                    # 実KS出货を優先（複数なら対比側キー）
                    raw = str(k_recs[i].get("出貨清單號") or "")
                    toks = split_ship_keys(raw)
                    if len(toks) == 1:
                        krow["K_出货清单号"] = toks[0]
                    elif k_ship_key and "/" not in k_ship_key:
                        krow["K_出货清单号"] = k_ship_key
                    elif toks:
                        # pick token matching soft pair
                        prefer = set(g.get("k_ships") or [])
                        hit = [t for t in toks if t in prefer]
                        krow["K_出货清单号"] = hit[0] if hit else toks[0]
                elif k_ship_key:
                    krow["K_出货清单号"] = k_ship_key
                row.update(krow)
            else:
                bk = blank_k()
                if k_ship_key and g["how"] not in ("caiwu_only",):
                    bk["K_出货清单号"] = k_ship_key if "/" not in k_ship_key else k_ship_key.split("/")[0]
                elif g["how"] == "caiwu_only" and c_ship_key:
                    bk["K_出货清单号"] = c_ship_key
                row.update(bk)
            out_rows.append(row)

    out = pd.DataFrame(out_rows)

    # sort: status priority by first row of each block — already in group order
    # re-order groups by status severity via stable sort on 状態 filled forward
    status_rank = {
        "CAIWUのみ（KS欠）": 0,
        "出荷番号R差（金額差）": 1,
        "注文番号不一致（金額差）": 2,
        "金額不一致": 3,
        "出荷番号R差（金額一致）": 4,
        "注文番号不一致（金額一致）": 5,
        "KSのみ（CAIWU欠）": 6,
        "一致": 7,
    }
    fill_status = []
    cur = ""
    for s in out["状態"]:
        if s:
            cur = s
        fill_status.append(cur)
    out["_rank"] = [status_rank.get(s, 9) for s in fill_status]
    out["_abs"] = pd.to_numeric(out["グループ金額差_KS-C"], errors="coerce").abs()
    # keep block integrity: assign block id
    block = -1
    bids = []
    for s in out["状態"]:
        if s:
            block += 1
        bids.append(block)
    out["_block"] = bids
    block_rank = out.groupby("_block")["_rank"].transform("min")
    block_abs = out.groupby("_block")["_abs"].transform("max")
    out["_rank"] = block_rank
    out["_abs"] = block_abs.fillna(0)
    out = out.sort_values(["_rank", "_abs", "_block"], ascending=[True, False, True]).drop(
        columns=["_rank", "_abs", "_block"]
    )

    stamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    args.out_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = (
        args.out_dir
        / f"caiwu_ks_line_detail_{args.from_month}_{args.to_month}_{stamp}.xlsx"
    )

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "明細対比"

    c_headers = [
        "品名", "客户名称", "项目名称", "订货单号", "sales", "产品名称", "产品尺寸",
        "数量", "单位", "每件", "每件净重量", "总净重量", "单价", "金额", "出货清单号",
    ]
    k_headers = [
        "品名", "客户名称", "单号", "日期", "本司订单编号", "项目名称", "销售人员",
        "出货清单号", "Description", "Qty", "UnitPrice", "Amount",
    ]
    meta = ["状態", "月", "突合", "対比出货", "グループ金額差", "グループC合計", "グループK合計"]
    # row1 section headers
    row1 = meta + ["CAIWU"] + [""] * (len(c_headers) - 1) + ["KS"] + [""] * (len(k_headers) - 1)
    row2 = ["", "", "", "出货清单号", "", "", ""] + c_headers + k_headers
    ws.append(row1)
    ws.append(row2)

    # merges
    meta_n = len(meta)
    c_n = len(c_headers)
    k_n = len(k_headers)
    ws.merge_cells(start_row=1, start_column=meta_n + 1, end_row=1, end_column=meta_n + c_n)
    ws.merge_cells(
        start_row=1,
        start_column=meta_n + c_n + 1,
        end_row=1,
        end_column=meta_n + c_n + k_n,
    )
    for col in range(1, meta_n + 1):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    caiwu_fill = PatternFill("solid", fgColor="2E75B6")
    ks_fill = PatternFill("solid", fgColor="548235")
    white = Font(color="FFFFFF", bold=True)
    for col in range(1, meta_n + c_n + k_n + 1):
        for r in (1, 2):
            cell = ws.cell(r, col)
            cell.font = white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col <= meta_n:
                cell.fill = header_fill
            elif col <= meta_n + c_n:
                cell.fill = caiwu_fill
            else:
                cell.fill = ks_fill

    fills = {
        "CAIWUのみ（KS欠）": PatternFill("solid", fgColor="FCE4D6"),
        "KSのみ（CAIWU欠）": PatternFill("solid", fgColor="DDEBF7"),
        "金額不一致": PatternFill("solid", fgColor="FFF2CC"),
        "注文番号不一致（金額差）": PatternFill("solid", fgColor="F8CBAD"),
        "注文番号不一致（金額一致）": PatternFill("solid", fgColor="E2EFDA"),
        "出荷番号R差（金額差）": PatternFill("solid", fgColor="F4B183"),
        "出荷番号R差（金額一致）": PatternFill("solid", fgColor="FCE4D6"),
        "一致": PatternFill("solid", fgColor="C6EFCE"),
    }

    data_cols = [
        "状態", "月", "突合", "対比出货清单号", "グループ金額差_KS-C", "グループC合計", "グループK合計",
        "C_品名", "C_客户名称", "C_项目名称", "C_订货单号", "C_sales", "C_产品名称", "C_产品尺寸",
        "C_数量", "C_单位", "C_每件", "C_每件净重量", "C_总净重量", "C_单价", "C_金额", "C_出货清单号",
        "K_品名", "K_客户名称", "K_单号", "K_日期", "K_本司订单编号", "K_项目名称", "K_销售人员",
        "K_出货清单号", "K_Description", "K_Qty", "K_UnitPrice", "K_Amount",
    ]

    num_cols = {
        "グループ金額差_KS-C", "グループC合計", "グループK合計",
        "C_数量", "C_每件", "C_每件净重量", "C_总净重量", "C_单价", "C_金额",
        "K_Qty", "K_UnitPrice", "K_Amount",
    }

    cur_status = ""
    for _, rec in out.iterrows():
        values = []
        for col in data_cols:
            v = rec.get(col, "")
            if v is None or (isinstance(v, float) and pd.isna(v)):
                v = ""
            values.append(v)
        ws.append(values)
        r = ws.max_row
        st = values[0] or cur_status
        if values[0]:
            cur_status = values[0]
        fill = fills.get(cur_status)
        if fill:
            for col in range(1, len(data_cols) + 1):
                ws.cell(r, col).fill = fill
        for ci, colname in enumerate(data_cols, 1):
            if colname in num_cols:
                cell = ws.cell(r, ci)
                if cell.value != "" and cell.value is not None:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = "#,##0.00"
                    except (TypeError, ValueError):
                        pass

    # summary
    ws2 = wb.create_sheet("サマリー")
    # group-level summary from first rows
    firsts = out[out["状態"].astype(str).str.len() > 0].copy()
    ws2.append(["状態", "グループ数", "C金額合計", "K金額合計", "差額合計"])
    for st in [
        "CAIWUのみ（KS欠）",
        "出荷番号R差（金額差）",
        "注文番号不一致（金額差）",
        "金額不一致",
        "出荷番号R差（金額一致）",
        "注文番号不一致（金額一致）",
        "KSのみ（CAIWU欠）",
        "一致",
    ]:
        sub = firsts[firsts["状態"] == st]
        if sub.empty:
            continue
        ws2.append(
            [
                st,
                len(sub),
                float(pd.to_numeric(sub["グループC合計"], errors="coerce").fillna(0).sum()),
                float(pd.to_numeric(sub["グループK合計"], errors="coerce").fillna(0).sum()),
                float(pd.to_numeric(sub["グループ金額差_KS-C"], errors="coerce").fillna(0).sum()),
            ]
        )
    ws2.append([])
    ws2.append(["明細行数", len(out)])
    ws2.append(["期間", f"{args.from_month} .. {args.to_month}"])
    ws2.append(["CAIWU", caiwu_path.name])
    ws2.append(["KS", ks_path.name])
    ws2.append(["注", "品名=classifyCaiwuProductName（佐近品名・dashboard同一）"])
    ws2.append(["注", "KS销售人员は貨值清單ヘッダから抽出"])

    # --- 品名×月 差（CAIWU正、KS−CAIWU）---
    months = sorted(
        {
            *c["sale_month"].dropna().astype(str).tolist(),
            *k["sale_month"].dropna().astype(str).tolist(),
        }
    )
    # pinming order from dashboard (fallback: alpha)
    pinming_order_path = SALES_DASHBOARD / "lib" / "caiwu-product-classify.ts"
    pinming_order: list[str] = []
    if pinming_order_path.exists():
        text = pinming_order_path.read_text(encoding="utf-8")
        # crude extract of quoted strings in CAIWU_PINMING_ORDER
        in_block = False
        for line in text.splitlines():
            if "CAIWU_PINMING_ORDER" in line:
                in_block = True
            if in_block:
                for m in re.findall(r'"([^"]+)"', line):
                    if m not in pinming_order:
                        pinming_order.append(m)
                if "]" in line and "as const" in line:
                    break
                if line.strip() == "];":
                    break
    if not pinming_order:
        pinming_order = sorted(set(c["品名"].tolist()) | set(k["品名"].tolist()))

    c_piv = (
        c.groupby(["品名", "sale_month"], dropna=False)["amount_hkd"]
        .sum()
        .unstack(fill_value=0)
    )
    k_piv = (
        k.groupby(["品名", "sale_month"], dropna=False)["Amount(HKD)"]
        .sum()
        .unstack(fill_value=0)
    )
    for m in months:
        if m not in c_piv.columns:
            c_piv[m] = 0.0
        if m not in k_piv.columns:
            k_piv[m] = 0.0
    c_piv = c_piv.reindex(columns=months, fill_value=0)
    k_piv = k_piv.reindex(columns=months, fill_value=0)

    all_pin = list(pinming_order)
    for p in sorted(set(c_piv.index.tolist()) | set(k_piv.index.tolist())):
        if p not in all_pin:
            all_pin.append(p)

    ws4 = wb.create_sheet("品名×月_差")
    header = ["品名"] + [f"C_{m}" for m in months] + ["C_合計"]
    header += [f"K_{m}" for m in months] + ["K_合計"]
    header += [f"差K-C_{m}" for m in months] + ["差_合計"]
    ws4.append(header)
    gap_fill_pos = PatternFill("solid", fgColor="FCE4D6")  # KS多い
    gap_fill_neg = PatternFill("solid", fgColor="DDEBF7")  # KS少ない
    for pin in all_pin:
        c_row = c_piv.loc[pin] if pin in c_piv.index else pd.Series(0, index=months)
        k_row = k_piv.loc[pin] if pin in k_piv.index else pd.Series(0, index=months)
        c_vals = [float(c_row.get(m, 0) or 0) for m in months]
        k_vals = [float(k_row.get(m, 0) or 0) for m in months]
        d_vals = [k_vals[i] - c_vals[i] for i in range(len(months))]
        c_tot = sum(c_vals)
        k_tot = sum(k_vals)
        d_tot = k_tot - c_tot
        if abs(c_tot) < 0.005 and abs(k_tot) < 0.005:
            continue
        ws4.append([pin] + c_vals + [c_tot] + k_vals + [k_tot] + d_vals + [d_tot])
        r = ws4.max_row
        for col in range(2, len(header) + 1):
            ws4.cell(r, col).number_format = "#,##0.00"
        # color total diff cell
        diff_col = len(header)
        cell = ws4.cell(r, diff_col)
        if d_tot > 0.5:
            cell.fill = gap_fill_pos
        elif d_tot < -0.5:
            cell.fill = gap_fill_neg

    # totals row
    c_month_tot = [float(c_piv[m].sum()) for m in months]
    k_month_tot = [float(k_piv[m].sum()) for m in months]
    d_month_tot = [k_month_tot[i] - c_month_tot[i] for i in range(len(months))]
    ws4.append(
        ["合計"]
        + c_month_tot
        + [sum(c_month_tot)]
        + k_month_tot
        + [sum(k_month_tot)]
        + d_month_tot
        + [sum(d_month_tot)]
    )
    r = ws4.max_row
    for col in range(2, len(header) + 1):
        ws4.cell(r, col).number_format = "#,##0.00"
        ws4.cell(r, col).font = Font(bold=True)
    ws4.freeze_panes = "B2"
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws4.max_row}"
    ws4.column_dimensions["A"].width = 28
    for i in range(2, len(header) + 1):
        ws4.column_dimensions[get_column_letter(i)].width = 12

    # 差の大きい品名トップ（合計）
    ws5 = wb.create_sheet("品名差トップ")
    ws5.append(["品名", "C_合計", "K_合計", "差_KS-CAIWU", "差|絶対|"])
    gap_rows = []
    for pin in all_pin:
        c_tot = float(c_piv.loc[pin].sum()) if pin in c_piv.index else 0.0
        k_tot = float(k_piv.loc[pin].sum()) if pin in k_piv.index else 0.0
        if abs(c_tot) < 0.005 and abs(k_tot) < 0.005:
            continue
        gap_rows.append((pin, c_tot, k_tot, k_tot - c_tot))
    gap_rows.sort(key=lambda x: abs(x[3]), reverse=True)
    for pin, c_tot, k_tot, d in gap_rows[:40]:
        ws5.append([pin, c_tot, k_tot, d, abs(d)])
        r = ws5.max_row
        for col in (2, 3, 4, 5):
            ws5.cell(r, col).number_format = "#,##0.00"
        if d > 0.5:
            ws5.cell(r, 4).fill = gap_fill_pos
        elif d < -0.5:
            ws5.cell(r, 4).fill = gap_fill_neg
    ws5.column_dimensions["A"].width = 28
    for col in "BCDE":
        ws5.column_dimensions[col].width = 14

    # KS欠 only sheet
    ws3 = wb.create_sheet("本当のKS欠")
    # simpler: filter rows belonging to those groups
    miss = out.copy()
    keep = []
    flag = False
    for _, r in miss.iterrows():
        if r["状態"] == "CAIWUのみ（KS欠）":
            flag = True
            keep.append(True)
        elif r["状態"] and r["状態"] != "CAIWUのみ（KS欠）":
            flag = False
            keep.append(False)
        else:
            keep.append(flag)
    miss = miss.loc[keep]
    ws3.append(meta + c_headers + k_headers)
    for _, rec in miss.iterrows():
        ws3.append([rec.get(c, "") for c in data_cols])

    widths = [18, 8, 10, 14, 12, 12, 12] + [14] * c_n + [14] * k_n
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 22)
    ws.auto_filter.ref = f"A2:{get_column_letter(len(data_cols))}{ws.max_row}"
    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30

    wb.save(out_xlsx)
    print(f"groups={len(groups)} detail_rows={len(out)}")
    print(firsts["状態"].value_counts().to_string())
    print("品名差トップ:")
    for pin, c_tot, k_tot, d in gap_rows[:10]:
        print(f"  {pin}: C={c_tot:,.0f} K={k_tot:,.0f} diff={d:,.0f}")
    print(f"xlsx: {out_xlsx}")


if __name__ == "__main__":
    main()
