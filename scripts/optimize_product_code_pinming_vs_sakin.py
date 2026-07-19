# -*- coding: utf-8 -*-
"""
佐近 Excel 品名×月 を正解として product_code -> 品名 を逆算する。
"""
from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

HIER = Path(
    r"E:\KSS-KSファイル集計\output\reports\KS_指定12項目抽出_分類階層_20260718_153724.xlsx"
)
SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
EXCL = Path(r"E:\sales-dashboard-2\data\ks-caiwu-excess-exclusions.json")
OUT = Path(r"E:\sales-dashboard-2\data\ks-product-code-pinming.json")
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
SKIP = re.compile(r"小\s*计|合\s*计|品\s*名|出口产品")
PAINT = re.compile(r"来料喷涂|喷涂加工|運費|运费")
ALIASES = {
    "铁天地结构骨槽": "铁槽",
    "铁制框架配件": "铁框架配件",
    "不锈钢框装饰板": "铝喷塑装饰板",
    "MIP石膏组合间隔610#": "石膏基高性能纤维板",
    "MIP石膏组合间隔1220#": "石膏基高性能纤维板",
    "12mm水泥纤维板-BOARD K": "12mm水泥纤维板-BOARD C",
    "12mm普通纸面石膏板": "普通纸面石膏板",
    "水泥纤维板-BOARD C": "12mm水泥纤维板-BOARD C",
}

# 佐近29品名 + 未分類（使わない）
PINMING = [
    "特造框架天花",
    "不锈钢框架天花",
    "特造铝制天花",
    "铝喷塑装饰板",
    "铝喷塑装饰配件",
    "铝方通",
    "铝铁组合",
    "KTA铝铁主骨",
    "KTA铝铁长副骨",
    "KTA铝铁短副骨",
    "铁喷塑装饰板",
    "铁装饰配件",
    "铁主龙骨",
    "铁长副龙骨",
    "铁短副龙骨",
    "铁L角",
    "铁框架配件",
    "不锈钢支架配件",
    "铁槽",
    "12mm水泥纤维板-BOARD C",
    "镀锌钢带",
    "9mm水泥纤维板-BOARD C",
    "普通纸面石膏板",
    "耐火纸面石膏板",
    "耐水纸面石膏板",
    "石膏基高性能纤维板",
    "岩棉-Y50\\60KG",
    "岩棉-B50\\100KG",
    "石膏基高性能纤维板(87\\112MM)",
    "石膏基高性能纤维板(N板)",
    "不锈钢框装饰板",
]


def load_sakin() -> dict[str, dict[str, float]]:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    out: dict[str, dict[str, float]] = {}
    for name in MONTHS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        out[name] = {}
        for r in range(4, (ws.max_row or 0) + 1):
            a = ws.cell(r, 1).value
            d = ws.cell(r, 4).value
            if not a:
                continue
            pin = ALIASES.get(str(a).strip(), str(a).strip())
            if SKIP.search(pin):
                continue
            if isinstance(d, (int, float)):
                out[name][pin] = out[name].get(pin, 0.0) + float(d)
    wb.close()
    return out


def seed_pinming(name: str, mid: str, desc: str) -> str:
    """line_detail + 佐近整合ルールから初期推定。"""
    name = (name or "").strip()
    mid = (mid or "").strip()
    blob = f"{name} {desc}"
    if PAINT.search(blob):
        return "EXCLUDE"
    n = re.sub(r"\s+", "", blob.lower())
    if "24t龙骨" in n or (name.startswith("24T") and "龙骨" in name):
        if re.search(r"短副|600mm|2尺|1200mm", n) and not re.search(r"3000|2400|10尺|主", n):
            return "铁短副龙骨"
        if re.search(r"长副|3000mm|2400mm|10尺", n):
            return "铁长副龙骨"
        return "铁主龙骨"
    if name in ("铁竖骨", "铁横骨", "铁中心骨") or "KTA15龙骨" in name:
        return "铁框架配件"
    if name in ("铁天地骨", "铁地槽", "铁C槽", "铝U型槽") or "天地骨" in name:
        return "铁槽"
    if "阔条天花" in blob or "格子天花" in blob or "铁双搭天花" in blob:
        return "特造框架天花"
    if name == "铝明架天花" or "铝明架天花" in blob:
        return "特造铝制天花"
    if "MK Board C" in blob or name.startswith("MK Board"):
        if re.search(r"9\s*mm|x\s*9mm", blob, re.I):
            return "9mm水泥纤维板-BOARD C"
        return "12mm水泥纤维板-BOARD C"
    if name in ("Typical Panel", "Non-Typical Panel"):
        return "铝喷塑装饰板"
    if "Stainless Steel" in blob or (name.startswith("Stainless") and "cladding" in blob.lower()):
        return "不锈钢框装饰板"
    if mid == "铁制天花板":
        return "特造框架天花"
    if mid == "不锈钢制天花板" or ("不锈钢" in name and "天花" in blob):
        return "不锈钢框架天花"
    if mid == "铁制框架配件":
        if re.search(r"天地骨|地槽", blob):
            return "铁槽"
        return "铁框架配件"
    if mid == "铝制天花板" or ("铝" in name and "天花" in blob):
        return "铝喷塑装饰板"
    if mid == "铝方通":
        return "铝方通"
    if mid == "镀锌钢带" or name == "镀锌带":
        return "镀锌钢带"
    if mid == "12MM普通纸面石膏板":
        return "普通纸面石膏板"
    if "Taishan HD GF" in blob or ("Taishan" in blob and re.search(r"\bN\b|N板", blob)):
        return "石膏基高性能纤维板(N板)"
    if "Taishan Regular" in blob:
        return "普通纸面石膏板"
    if "Fireproof" in blob and "Gypsum" in blob:
        return "耐火纸面石膏板"
    if "Moisture" in blob and "Gypsum" in blob:
        return "耐水纸面石膏板"
    if "Taishan" in blob or "MIP" in blob:
        return "石膏基高性能纤维板"
    if "Welltone" in blob and "100" in blob:
        return "岩棉-B50\\100KG"
    if "Welltone" in blob:
        return "岩棉-Y50\\60KG"
    if mid.startswith("铝"):
        return "铝喷塑装饰配件" if any(x in mid for x in ("角", "槽", "风咀", "支架", "配件")) else "铝喷塑装饰板"
    return "铁框架配件" if mid == "龙骨(桐井）" else ""


def main() -> None:
    sakin = load_sakin()
    target: dict[tuple[str, str], float] = {}
    for m in MONTHS:
        for pin, amt in sakin.get(m, {}).items():
            target[(m, pin)] = amt

    k = pd.read_excel(HIER, sheet_name="明细")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["month"] = k["日期"].dt.strftime("%Y-%m")
    k = k[k["month"].isin(MONTHS)].copy()
    if EXCL.exists():
        data = json.loads(EXCL.read_text(encoding="utf-8"))
        k = k[~k.index.isin(set(data.get("excluded_row_indices", [])))]
    k = k[~k["description"].astype(str).str.contains(PAINT, na=False)]

    # product_code -> month -> amount
    code_month: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    code_seed: dict[str, str] = {}
    code_info: dict[str, dict] = {}
    for _, r in k.iterrows():
        code = str(r.get("产品代码") or "").strip()
        if not code:
            continue
        m = str(r["month"])
        code_month[code][m] += float(r["Amount(HKD)"])
        if code not in code_seed:
            code_seed[code] = seed_pinming(
                str(r.get("品名") or ""),
                str(r.get("材料中分类") or ""),
                str(r.get("description") or ""),
            )
            code_info[code] = {
                "name": str(r.get("品名") or ""),
                "mid": str(r.get("材料中分类") or ""),
            }

    # current aggregate from seeds
    cur: dict[tuple[str, str], float] = defaultdict(float)
    for code, months in code_month.items():
        seed = code_seed.get(code) or "铝喷塑装饰板"
        if seed == "EXCLUDE":
            continue
        seed = ALIASES.get(seed, seed)
        if seed not in PINMING:
            seed = "铝喷塑装饰板"
        for m, amt in months.items():
            cur[(m, seed)] += amt

    def total_error(assign: dict[str, str]) -> float:
        agg: dict[tuple[str, str], float] = defaultdict(float)
        for code, months in code_month.items():
            pin = assign.get(code, "铝喷塑装饰板")
            pin = ALIASES.get(pin, pin)
            if pin not in PINMING:
                continue
            for m, amt in months.items():
                agg[(m, pin)] += amt
        err = 0.0
        for key, t in target.items():
            err += abs(agg.get(key, 0.0) - t)
        return err

    assign = {}
    for code, months in code_month.items():
        seed = code_seed.get(code) or ""
        seed = ALIASES.get(seed, seed)
        assign[code] = seed if seed in PINMING else "铝喷塑装饰板"

    # greedy improve: try reassign each code to best pinming
    codes = sorted(code_month.keys(), key=lambda c: sum(code_month[c].values()), reverse=True)
    best_err = total_error(assign)
    improved = True
    rounds = 0
    while improved and rounds < 8:
        improved = False
        rounds += 1
        for code in codes:
            base = assign[code]
            best_pin = base
            best_local = best_err
            for pin in PINMING:
                if pin == base:
                    continue
                assign[code] = pin
                e = total_error(assign)
                if e < best_local - 100:
                    best_local = e
                    best_pin = pin
            assign[code] = best_pin
            if best_local < best_err - 100:
                best_err = best_local
                improved = True

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "source": "optimized_from_sakin_excel",
        "hierarchy": HIER.name,
        "count": len(assign),
        "errorHkd": round(best_err, 2),
        "byCode": {c: ALIASES.get(p, p) for c, p in assign.items()},
        "seedInfo": code_info,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(assign)} codes error={best_err:,.0f} -> {OUT}")


if __name__ == "__main__":
    main()
