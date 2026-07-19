# -*- coding: utf-8 -*-
"""Validate KS classification vs 佐近 Excel after classifier fix."""
from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
KS_XLSX = Path(
    r"E:\KSS-KSファイル集計\output\reports\KS_指定12項目抽出_分類階層_20260718_153724.xlsx"
)
EXCL = Path(r"E:\sales-dashboard-2\data\ks-caiwu-excess-exclusions.json")
SALES = Path(r"E:\sales-dashboard-2")
BATCH = SALES / "scripts" / "classify-pinming-batch.ts"
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
PAINT = re.compile(r"来料喷涂|喷涂加工|運費|运费")
ALIASES = {
    "铁天地结构骨槽": "铁槽",
    "铁制框架配件": "铁框架配件",
    "不锈钢框装饰板": "铝喷塑装饰板",
    "MIP石膏组合间隔610#": "石膏基高性能纤维板",
    "MIP石膏组合间隔1220#": "石膏基高性能纤维板",
    "12mm水泥纤维板-BOARD K": "12mm水泥纤维板-BOARD C",
    "12mm普通纸面石膏板": "普通纸面石膏板",
    "水泥纤维板-BOARD C": "12mm水泥纤维板-BOARD C",
}


def alias_pinming(name: str) -> str:
    return ALIASES.get(name, name)


def sakin_by_pinming() -> pd.DataFrame:
    wb = load_workbook(SAKIN, data_only=True)
    rows = []
    for name in MONTHS:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for r in range(1, (ws.max_row or 0) + 1):
            pin = ws.cell(r, 1).value
            amt = ws.cell(r, 4).value
            if not pin or not str(pin).strip():
                continue
            pin_s = re.sub(r"\s+", " ", str(pin).strip())
            if re.search(r"小\s*计|合\s*计|品\s*名|出口产品|明细表", pin_s):
                continue
            try:
                amt_f = float(amt or 0)
            except (TypeError, ValueError):
                amt_f = 0.0
            rows.append({"sale_month": name, "品名": pin_s, "佐近_HKD": amt_f})
    wb.close()
    return pd.DataFrame(rows)


def classify_batch(req: list[dict]) -> dict[str, str]:
    tmp = Path(r"E:\factory_monitoring_system\logs\_tmp_pinming")
    tmp.mkdir(exist_ok=True)
    inp, outp = tmp / "val_in.json", tmp / "val_out.json"
    inp.write_text(json.dumps(req, ensure_ascii=False), encoding="utf-8")
    exe = "npx.cmd" if os.name == "nt" else "npx"
    proc = subprocess.run(
        [exe, "--yes", "tsx", str(BATCH), str(inp), str(outp)],
        cwd=str(SALES),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    if proc.returncode != 0:
        raise SystemExit(proc.stderr or proc.stdout)
    return {r["id"]: r["pinming"] for r in json.loads(outp.read_text(encoding="utf-8"))}


def main() -> None:
    sakin = sakin_by_pinming()
    k = pd.read_excel(KS_XLSX, sheet_name="明细")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["sale_month"] = k["日期"].dt.strftime("%Y-%m")
    k = k[k["sale_month"].isin(MONTHS)].copy()

    excl_idx = set()
    if EXCL.exists():
        data = json.loads(EXCL.read_text(encoding="utf-8"))
        excl_idx.update(data.get("excluded_row_indices", []))
    k = k[~k.index.isin(excl_idx)]
    k = k[~k["description"].astype(str).str.contains(PAINT, na=False)]

    req = []
    for i, r in k.iterrows():
        desc = r.get("description")
        pname = r.get("品名")
        if pd.isna(desc):
            desc = None
        if pd.isna(pname):
            pname = None
        req.append(
            {
                "id": str(i),
                "product_name": None if pd.isna(r.get("品名")) else r.get("品名"),
                "description": None if pd.isna(r.get("description")) else r.get("description"),
                "mid_category": None if pd.isna(r.get("材料中分类")) else r.get("材料中分类"),
                "product_code": None if pd.isna(r.get("产品代码")) else r.get("产品代码"),
            }
        )
    mp = classify_batch(req)
    k["品名"] = [mp[str(i)] for i in k.index]

    ks = k.groupby(["sale_month", "品名"])["Amount(HKD)"].sum().reset_index()
    ks = ks.rename(columns={"Amount(HKD)": "KS_HKD"})

    m = sakin.merge(ks, on=["sale_month", "品名"], how="outer").fillna(0)
    m["差"] = m["KS_HKD"] - m["佐近_HKD"]

    m["品名_統合"] = m["品名"].map(alias_pinming)
    mu = m.groupby(["sale_month", "品名_統合"])[["佐近_HKD", "KS_HKD"]].sum().reset_index()
    mu["差"] = mu["KS_HKD"] - mu["佐近_HKD"]

    print("=== 月合計 ===")
    for month in MONTHS:
        sm = m.loc[m["sale_month"] == month, "佐近_HKD"].sum()
        km = m.loc[m["sale_month"] == month, "KS_HKD"].sum()
        print(f"{month} 佐近={sm:,.0f} KS={km:,.0f} 差={km-sm:+,.0f}")

    print("\n=== 品名差トップ (1-6, 統合後) ===")
    tot = mu.groupby("品名_統合")[["佐近_HKD", "KS_HKD"]].sum()
    tot["差"] = tot["KS_HKD"] - tot["佐近_HKD"]
    tot = tot.sort_values("差", key=lambda s: s.abs(), ascending=False)
    print(tot.head(15).to_string())

    print("\n=== 品名差トップ (raw) ===")
    tot_raw = m.groupby("品名")[["佐近_HKD", "KS_HKD"]].sum()
    tot_raw["差"] = tot_raw["KS_HKD"] - tot_raw["佐近_HKD"]
    print(tot_raw.sort_values("差", key=lambda s: s.abs(), ascending=False).head(8).to_string())

    out = Path(r"C:\Users\Satoshi\Downloads") / "KS分類_佐近比較_検証.xlsx"
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        mu.sort_values(["sale_month", "品名_統合"]).to_excel(w, index=False, sheet_name="品名×月_統合")
        m.sort_values(["sale_month", "品名"]).to_excel(w, index=False, sheet_name="品名×月_raw")
        tot.reset_index().to_excel(w, index=False, sheet_name="品名合計_統合")
    print(f"\nwrote {out}")


if __name__ == "__main__":
    main()
