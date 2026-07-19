# -*- coding: utf-8 -*-
"""Deploy前チェック: 佐近Excel品名がダッシュボードに出るか + 主要品名×月."""
from __future__ import annotations

import json
import os
import re
import subprocess
from pathlib import Path

import openpyxl
import pandas as pd

SAKIN = Path(r"C:\Users\Satoshi\Downloads\销售收入-2026（佐近先生）(1).xlsx")
HIER = Path(
    r"E:\KSS-KSファイル集計\output\reports\KS_指定12項目抽出_分類階層_20260718_153724.xlsx"
)
EXCL = Path(r"E:\sales-dashboard-2\data\ks-caiwu-excess-exclusions.json")
SALES = Path(r"E:\sales-dashboard-2")
BATCH = SALES / "scripts" / "classify-pinming-batch.ts"
TS = SALES / "lib" / "caiwu-product-classify.ts"
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
SKIP = re.compile(r"小\s*计|合\s*计|品\s*名|出口产品")
PAINT = re.compile(r"来料喷涂|喷涂加工|運費|运费")

# Excel-only rows merged in display (allowed)
DISPLAY_ALIASES = {
    "铁天地结构骨槽",
    "铁制框架配件",
    "不锈钢框装饰板",
    "MIP石膏组合间隔610#",
    "MIP石膏组合间隔1220#",
    "12mm水泥纤维板-BOARD K",
    "12mm普通纸面石膏板",
    "水泥纤维板-BOARD C",
    "石膏基高性能纤维板(墙板)",
}


def load_dashboard_order() -> list[str]:
    text = TS.read_text(encoding="utf-8")
    m = re.search(r"CAIWU_PINMING_ORDER = \[(.*?)\]", text, re.S)
    if not m:
        return []
    return [
        x.strip().strip('"')
        for x in m.group(1).split(",")
        if x.strip() and x.strip('"') != "未分類"
    ]


def load_sakin_names() -> set[str]:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    names: set[str] = set()
    for sh in wb.sheetnames:
        if not re.match(r"2026-\d{2}", sh):
            continue
        ws = wb[sh]
        for r in range(4, (ws.max_row or 0) + 1):
            a = ws.cell(r, 1).value
            if not a:
                continue
            a = str(a).strip()
            if SKIP.search(a):
                continue
            names.add(a)
    wb.close()
    return names


def load_sakin_month(name: str, month: str) -> float:
    wb = openpyxl.load_workbook(SAKIN, data_only=True)
    ws = wb[month]
    total = 0.0
    for r in range(4, (ws.max_row or 0) + 1):
        a = ws.cell(r, 1).value
        d = ws.cell(r, 4).value
        if a and str(a).strip() == name and isinstance(d, (int, float)):
            total += float(d)
    wb.close()
    return total


def classify_ks() -> pd.DataFrame:
    k = pd.read_excel(HIER, sheet_name="明细")
    k["Amount(HKD)"] = pd.to_numeric(k["Amount(HKD)"], errors="coerce").fillna(0)
    k["日期"] = pd.to_datetime(k["日期"], errors="coerce")
    k["month"] = k["日期"].dt.strftime("%Y-%m")
    k = k[k["month"].isin(MONTHS)].copy()
    k = k[~k["description"].astype(str).str.contains(PAINT, na=False)]
    if EXCL.exists():
        data = json.loads(EXCL.read_text(encoding="utf-8"))
        k = k[~k.index.isin(set(data.get("excluded_row_indices", [])))]
        keys = {
            (
                str(l.get("ks_doc_no", "")).upper(),
                str(l.get("delivery_list_no", "")).upper(),
                str(l.get("description", ""))[:80],
                round(float(l.get("amount_hkd", 0)), 2),
            )
            for l in data.get("lines", [])
        }
        k = k[
            ~k.apply(
                lambda r: (
                    str(r.get("單號", "")).upper(),
                    str(r.get("出貨清單號", "")).upper(),
                    str(r.get("description", ""))[:80],
                    round(float(r.get("Amount(HKD)", 0)), 2),
                )
                in keys,
                axis=1,
            )
        ]
    req = []
    for i, r in k.iterrows():
        req.append(
            {
                "id": str(i),
                "product_name": None if pd.isna(r.get("品名")) else r.get("品名"),
                "description": None if pd.isna(r.get("description")) else r.get("description"),
                "mid_category": None if pd.isna(r.get("材料中分类")) else r.get("材料中分类"),
                "product_code": None if pd.isna(r.get("产品代码")) else r.get("产品代码"),
            }
        )
    tmp = Path(r"E:\factory_monitoring_system\logs\_tmp_checklist")
    tmp.mkdir(exist_ok=True)
    inp, outp = tmp / "in.json", tmp / "out.json"
    inp.write_text(json.dumps(req, ensure_ascii=False), encoding="utf-8")
    exe = "npx.cmd" if os.name == "nt" else "npx"
    proc = subprocess.run(
        [exe, "--yes", "tsx", str(BATCH), str(inp), str(outp)],
        cwd=str(SALES),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        shell=os.name == "nt",
    )
    if proc.returncode != 0:
        raise SystemExit(proc.stderr or proc.stdout)
    mp = {x["id"]: x["pinming"] for x in json.loads(outp.read_text(encoding="utf-8"))}
    k["品名"] = [mp[str(i)] for i in k.index]
    return k


def main() -> None:
    order = load_dashboard_order()
    sakin_names = load_sakin_names()
    missing = sorted(
        n for n in sakin_names if n not in order and n not in DISPLAY_ALIASES
    )
    print("=== CHECK 1: Sakin Excel pinming missing from dashboard order ===")
    if missing:
        for n in missing:
            print(f"  MISSING: {n}")
    else:
        print("  OK (all Sakin rows in dashboard or known alias)")

    print(f"\n  不锈钢框架天花 in order: {'不锈钢框架天花' in order}")

    k = classify_ks()
    must = ["不锈钢框架天花", "特造框架天花", "特造铝制天花", "铝喷塑装饰板", "铁框架配件", "铁槽"]
    print("\n=== CHECK 2: pinming x month (Sakin vs KS) ===")
    for name in must:
        print(f"\n-- {name} --")
        for month in MONTHS:
            sk = load_sakin_month(name, month)
            ks = float(k[(k["month"] == month) & (k["品名"] == name)]["Amount(HKD)"].sum())
            if sk or ks:
                print(f"  {month} Sakin {sk:,.0f}  KS {ks:,.0f}  diff {ks - sk:+,.0f}")


if __name__ == "__main__":
    main()
